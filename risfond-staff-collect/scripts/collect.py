# -*- coding: utf-8 -*-
"""
锐仕方达内部系统 人选信息收集脚本
流程：列表页(不关闭) -> 新标签打开详情 -> 查看联系方式 -> 确认(6R币) -> 抓取 -> 关标签回列表 -> 翻页
用法：python collect.py <目标人数>
信号：等待 staff-collect\GO3 文件
增量保存 records.json，跑完生成 Excel。
"""
import sys
import io
import json
import time
import random
import os
from pathlib import Path
from playwright.sync_api import sync_playwright

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stdout.reconfigure(line_buffering=True)

# ============ 可配置区（分发给同事时只需改这里）============
# 工作目录：存放 records.json / Excel / 浏览器登录配置
BASE = Path(os.environ.get("RISFOND_DIR", r"D:\hermes猎头\staff-collect"))
# 人选列表页网址
URL = os.environ.get("RISFOND_URL", "https://staff.risfond.com/resume/nsearchresume?keywords=")
# 每人之间的随机停留秒数（模拟真人节奏）
SLEEP_PER_PERSON = (3, 10)
# 翻页随机停留秒数
SLEEP_PER_PAGE = (3, 8)
# ========================================================

PROFILE = str(BASE / "browser-profile")
GO3 = BASE / "GO3"
RECORDS = BASE / "records.json"
EXCEL = BASE / "人选信息收集.xlsx"
TARGET = int(sys.argv[1]) if len(sys.argv) > 1 else 100

GO3.unlink(missing_ok=True)


def load_records():
    if RECORDS.exists():
        try:
            return json.loads(RECORDS.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []


def save_records(recs):
    RECORDS.write_text(json.dumps(recs, ensure_ascii=False, indent=1), encoding="utf-8")


def close_popups(pg):
    for sel in [".workSummaryOfYesterday__title_right img",
                ".ZhiNengXiaoDa_hea_close",
                ".el-dialog__headerbtn", ".el-icon-close", ".modal .close",
                ".tel-modal-close", "button.close"]:
        try:
            for el in pg.locator(sel).all():
                if el.is_visible(timeout=300):
                    el.click()
                    pg.wait_for_timeout(300)
        except Exception:
            pass


def wait_phone_ready(detail, timeout=12):
    """等电话元素出现且innerText非空"""
    try:
        detail.wait_for_function(
            "() => { const els = document.querySelectorAll('#in_phonenumber'); "
            "return els.length && Array.from(els).some(e => (e.innerText || '').trim()); }",
            timeout=timeout * 1000,
        )
        return True
    except Exception:
        return False


def read_phones(detail):
    # 注意：all_inner_texts() 不接受 timeout 参数，传了会抛 TypeError 被吞掉返回空！
    try:
        parts = detail.locator("#in_phonenumber").all_inner_texts()
        return [x.strip() for x in parts if x.strip()]
    except Exception:
        return []


def read_basic(detail, rec):
    try:
        rec["职位"] = detail.locator("#in_currentjobtitle").inner_text(timeout=5000).strip()
    except Exception:
        pass
    try:
        rec["目前公司"] = detail.locator("#in_currentcompany").inner_text(timeout=5000).strip()
    except Exception:
        pass


def find_visible_btn(detail):
    btns = detail.locator("text=查看联系方式")
    try:
        n = btns.count()
    except Exception:
        return None
    for i in range(n):
        b = btns.nth(i)
        try:
            if b.is_visible(timeout=600):
                return b
        except Exception:
            continue
    return None


def collect_one(ctx, page, cand, idx, total):
    cid = cand["id"]
    name = cand["name"]
    print(f"[{idx}/{total}] 收集：{name} ...")
    rec = {"id": cid, "姓名": name, "电话": "", "职位": "", "目前公司": "", "状态": ""}
    try:
        with ctx.expect_page(timeout=30000) as np:
            page.locator(f'a[href*="viewresume?id={cid}"]').first.click()
        detail = np.value
        detail.wait_for_load_state("domcontentloaded", timeout=25000)
        close_popups(detail)

        # 等电话非空 + 固定稳定时间
        wait_phone_ready(detail, timeout=12)
        detail.wait_for_timeout(2000)

        phones = read_phones(detail)
        clean = [x for x in phones if "*" not in x]
        masked = [x for x in phones if "*" in x]

        if clean:
            # 已解锁：直接读
            rec["电话"] = " / ".join(clean)
            read_basic(detail, rec)
            rec["状态"] = "成功"
            print("    (已解锁，免R币直接读取)")
        elif masked:
            # 未解锁：花币流程
            btn = find_visible_btn(detail)
            if btn is not None:
                btn.click(timeout=10000)
                detail.wait_for_timeout(random.uniform(800, 1500))
                detail.locator("button.btn_Save").first.click(timeout=10000)
                try:
                    detail.wait_for_function(
                        "() => { const els = document.querySelectorAll('#in_phonenumber'); "
                        "return els.length && Array.from(els).some(e => (e.innerText || '').trim() && !e.innerText.includes('*')); }",
                        timeout=20000,
                    )
                except Exception:
                    pass
                rec["电话"] = " / ".join(read_phones(detail))
                read_basic(detail, rec)
                rec["状态"] = "成功"
            else:
                detail.wait_for_timeout(3000)
                phones = read_phones(detail)
                clean2 = [x for x in phones if "*" not in x]
                if clean2:
                    rec["电话"] = " / ".join(clean2)
                    read_basic(detail, rec)
                    rec["状态"] = "成功"
                    print("    (延迟解锁，免R币直接读取)")
                else:
                    rec["状态"] = "失败: 电话脱敏且无按钮"
        else:
            # 电话空：再等一次
            detail.wait_for_timeout(3000)
            phones = read_phones(detail)
            clean = [x for x in phones if "*" not in x]
            if clean:
                rec["电话"] = " / ".join(clean)
                read_basic(detail, rec)
                rec["状态"] = "成功"
                print("    (延迟填充，免R币直接读取)")
            else:
                rec["状态"] = "失败: 电话为空"

        try:
            detail.close()
        except Exception:
            pass
        if rec["状态"] == "成功":
            print(f"    ✓ {name} | {rec['电话']} | {rec['职位']} | {rec['目前公司']}")
        return rec
    except Exception as e:
        rec["状态"] = f"失败: {str(e)[:80]}"
        print(f"    ✗ {name} 失败: {str(e)[:120]}")
        try:
            for pg in ctx.pages[1:]:
                pg.close()
        except Exception:
            pass
    return rec


def main():
    records = load_records()
    done_ids = {r["id"] for r in records if r.get("状态") == "成功"}
    print(f"已有成功记录 {len(done_ids)} 条，本批目标 {TARGET} 条")

    with sync_playwright() as p:
        ctx = p.chromium.launch_persistent_context(
            PROFILE, headless=False, no_viewport=True,
            args=["--start-maximized"], locale="zh-CN",
        )
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
        page.goto(URL, wait_until="domcontentloaded", timeout=120000)
        print(">>> 请在浏览器中登录并搜索出人选列表，完成后由agent创建GO3信号 <<<")
        while not GO3.exists():
            time.sleep(2)
        print("收到GO3信号，开始收集...")
        page.wait_for_timeout(2000)
        close_popups(page)

        page_no = 1
        while len(done_ids) < TARGET:
            raw = page.eval_on_selector_all(
                'a[href*="viewresume?id="]',
                """els => els.map(e => {
                    const m = (e.href||'').match(/id=(\\d+)/);
                    const t = (e.innerText||'').trim();
                    const okName = t && t.length >= 2 && t.length <= 6
                        && !/查看详情|留言|加微|给人选|系统简历/.test(t);
                    return (m && okName) ? {id: m[1], name: t} : null;
                }).filter(x => x)""",
            )
            seen = set()
            cands = []
            for r in raw:
                if r["id"] not in seen:
                    seen.add(r["id"])
                    cands.append(r)
            cands = [c for c in cands if c["id"] not in done_ids]
            print(f"--- 第{page_no}页：{len(cands)} 位待收集 ---")

            for cand in cands:
                if len(done_ids) >= TARGET:
                    break
                rec = collect_one(ctx, page, cand, len(done_ids) + 1, TARGET)
                records.append(rec)
                save_records(records)
                if rec["状态"] == "成功":
                    done_ids.add(rec["id"])
                time.sleep(random.uniform(*SLEEP_PER_PERSON))
                close_popups(page)

            if len(done_ids) >= TARGET:
                break
            flipped = False
            for attempt in range(3):
                try:
                    close_popups(page)
                    nxt = page.locator(".jqPager >> text=下一页").first
                    if not nxt.is_visible(timeout=3000):
                        print("没有下一页了，结束")
                        flipped = None
                        break
                    nxt.click()
                    page.wait_for_timeout(random.uniform(2000, 3000))
                    page.locator('a[href*="viewresume?id="]').first.wait_for(timeout=30000)
                    page_no += 1
                    flipped = True
                    break
                except Exception as e:
                    print(f"    翻页尝试{attempt+1}失败: {str(e)[:60]}，重试...")
                    page.wait_for_timeout(3000)
            if flipped is None:
                break
            if not flipped:
                print("翻页多次失败，结束")
                break
            time.sleep(random.uniform(*SLEEP_PER_PAGE))

        ctx.close()

    ok = [r for r in records if r.get("状态") == "成功"]
    if ok:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "人选信息"
        headers = ["姓名", "电话", "职位", "目前公司"]
        ws.append(headers)
        fill = PatternFill("solid", fgColor="F28C28")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for r in ok:
            ws.append([r["姓名"], r["电话"], r["职位"], r["目前公司"]])
        for col, w in zip("ABCD", [14, 20, 28, 40]):
            ws.column_dimensions[col].width = w
        wb.save(str(EXCEL))
        print(f"Excel 已生成：{EXCEL}（{len(ok)} 条）")
    print(f"收集结束：成功 {len(ok)} / 目标 {TARGET}")


if __name__ == "__main__":
    main()
